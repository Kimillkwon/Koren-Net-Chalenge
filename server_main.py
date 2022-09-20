from socket import *
from os.path import exists
import sys
import os
import time
import torch
from tensorflow.python.keras.utils import np_utils
from tensorflow.python.keras.models import Sequential
from tensorflow.python.keras.layers import Dense, Activation
from tensorflow.python.keras.models import load_model
from keras.utils import np_utils
from numpy import argmax
import tensorflow as tf
from torch import nn
import torch.nn.functional as F
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
import gluonnlp as nlp
import numpy as np
from tqdm import tqdm, tqdm_notebook
from kobert.utils import get_tokenizer
from kobert.pytorch_kobert import get_pytorch_kobert_model
from transformers import AdamW
from transformers.optimization import get_cosine_schedule_with_warmup
import os
from sklearn.model_selection import train_test_split
import pandas as pd
import cv2
import gc
gc.collect()
torch.cuda.empty_cache()

end = "false"
model = load_model('model.h5',compile=False)

class BERTDataset(Dataset):
    def __init__(self, dataset, sent_idx, label_idx, bert_tokenizer, max_len,
                 pad, pair):
        transform = nlp.data.BERTSentenceTransform(
            bert_tokenizer, max_seq_length=max_len, pad=pad, pair=pair)

        self.sentences = [transform([i[sent_idx]]) for i in dataset]
        self.labels = [np.int32(i[label_idx]) for i in dataset]

    def __getitem__(self, i):
        return (self.sentences[i] + (self.labels[i], ))

    def __len__(self):
        return (len(self.labels))

class BERTClassifier(nn.Module):
    def __init__(self,
                 bert,
                 hidden_size = 768,
                 num_classes=7,
                 dr_rate=None,
                 params=None):
        super(BERTClassifier, self).__init__()
        self.bert = bert
        self.dr_rate = dr_rate

        self.classifier = nn.Linear(hidden_size , num_classes)
        if dr_rate:
            self.dropout = nn.Dropout(p=dr_rate)

    def gen_attention_mask(self, token_ids, valid_length):
        attention_mask = torch.zeros_like(token_ids)
        for i, v in enumerate(valid_length):
            attention_mask[i][:v] = 1
        return attention_mask.float()

    def forward(self, token_ids, valid_length, segment_ids):
        attention_mask = self.gen_attention_mask(token_ids, valid_length)

        _, pooler = self.bert(input_ids = token_ids, token_type_ids = segment_ids.long(), attention_mask = attention_mask.float().to(token_ids.device))
        if self.dr_rate:
            out = self.dropout(pooler)
        return self.classifier(out)


def txt():
        serverSock = socket(AF_INET, SOCK_STREAM)
        serverSock.setsockopt(SOL_SOCKET, SO_REUSEADDR, 1)
        serverSock.bind(('10.246.246.65', 15000))
        serverSock.listen(1)

        connectionSock, addr = serverSock.accept()
        filename = connectionSock.recv(1024)
        print('받은 데이터 : ', filename.decode('utf-8'))

        nowdir = os.getcwd()
        data = connectionSock.recv(1024)
        with open(nowdir+"/"+filename.decode('utf-8'), 'w') as f:
                f.write(data.decode('utf-8'))
        f.close()
        serverSock.close()
        return data.decode('utf-8')

def img():
        serverSock = socket(AF_INET, SOCK_STREAM)
        serverSock.setsockopt(SOL_SOCKET, SO_REUSEADDR, 1)
        serverSock.bind(('10.246.246.65', 10000))
        serverSock.listen(1)

        connectionSock, addr = serverSock.accept()
        print(str(addr),'에서 접속했습니다')

        filename = connectionSock.recv(1024)
        print('받은 데이터 : ', filename.decode('utf-8'))

        data_transferred = 0
        data = connectionSock.recv(1024)
        nowdir = os.getcwd()
        with open(nowdir+"/"+filename.decode('utf-8'), 'wb') as f:
                try:
                        while data:
                                f.write(data)
                                data_transferred += len(data)
                                data = connectionSock.recv(1024)
                except Exception as ex:
                        print(ex)
        #print('파일 %s 받기 완료. 전송량 %d' %(filename, data_transferred))
        serverSock.close()

def printtxt(num):
        filename = "output"+str(num)+".txt"
        ftext = open(filename,'r')
        line = ftext.readline()
        ftext.close()
        print(line)


def recvend():
        global end
        serverSock = socket(AF_INET, SOCK_STREAM)
        serverSock.setsockopt(SOL_SOCKET, SO_REUSEADDR, 1)
        serverSock.bind(('10.246.246.65', 13000))
        serverSock.listen(1)

        connectionSock, addr = serverSock.accept()
        endmsg = connectionSock.recv(1024)
        end=endmsg.decode('utf-8')
        serverSock.close()

def sendEmotion(emotion):
    clientSock = socket(AF_INET, SOCK_STREAM)
    clientSock.connect(('203.255.253.151', 8080))

    clientSock.sendall(emotion.encode('utf-8'))
    time.sleep(1)
    clientSock.close()

def predict(predict_sentence):
        max_len = 64
        batch_size = 64 
        warmup_ratio = 0.1
        num_epochs = 10
        max_grad_norm = 1
        log_interval = 200
        learning_rate =  5e-5
        emotion = ''

        data = [predict_sentence, '0']
        dataset_another = [data]

        another_test = BERTDataset(dataset_another, 0, 1, tok, max_len, True, False)
        test_dataloader = torch.utils.data.DataLoader(another_test, batch_size=batch_size, num_workers=5)

#        model1.eval()

        for batch_id, (token_ids, valid_length, segment_ids, label) in enumerate(test_dataloader):
            token_ids = token_ids.long().to(device)
            segment_ids = segment_ids.long().to(device)

            valid_length= valid_length
            label = label.long().to(device)

            out = model1(token_ids, valid_length, segment_ids)


            test_eval=[]
            for i in out:
                logits=i
                logits = logits.detach().cpu().numpy()

                if np.argmax(logits) == 0:
                    test_eval.append("공포가")
                    emotions = -1
                elif np.argmax(logits) == 1:
                    test_eval.append("놀람이")
                    emotions = -1
                elif np.argmax(logits) == 2:
                    test_eval.append("분노가")
                    emotions = -1
                elif np.argmax(logits) == 3:
                    test_eval.append("슬픔이")
                    emotions = -1
                elif np.argmax(logits) == 4:
                    test_eval.append("중립이")
                    emotions = 0
                elif np.argmax(logits) == 5:
                    test_eval.append("행복이")
                    emotions = 1
                elif np.argmax(logits) == 6:
                    test_eval.append("혐오가")
                    emotions = -1

            print("음성인식엔진: " + test_eval[0] + " 느껴집니다.")
            return emotions

def getCNN(num):
    image_w = 150
    image_h = 150
    X=[]
    base_dir = '/home2/siso/kik/'
    file_name =  'output'+str(num)+'.png'
    final_dir = os.path.join(base_dir+file_name)
    stream = open(final_dir, "rb")
    bytes = bytearray(stream.read())
    nparray = np.asarray(bytes, dtype=np.uint8)
    img = cv2.imdecode(nparray,cv2.IMREAD_UNCHANGED)
    img = cv2.resize(img, None, fx=image_w/img.shape[1], fy=image_h/img.shape[0])
    X.append(img/256)
    X = np.array(X)
    Max = np.amax(X); Min = np.amin(X); X = (X-Min) / (Max-Min)

    # Predict
    xhat_idx = np.random.choice(X.shape[0],1)
    xhat = X[xhat_idx]
    yhat = np.argmax(model.predict(X), axis=-1)

    # Mapping
    if yhat[0] == 0:
        emotion = '행복'
        emotions = 0.3 
    elif yhat[0] ==  1:
        emotion = '우울'
        emotions = -0.3
    elif yhat[0] == 2:
        emotion = '우울'
        emotions = -0.3
    elif yhat[0] == 3:
        emotion = '우울'
        emotions = -0.3
    elif yhat[0] == 4:
        emotion = '우울'
        emotions = -0.3
    elif yhat[0] == 5:
        emotion = '우울'
        emotions = -0.3

    #print(yhat.shape, type(yhat), len(yhat), yhat)
    stream.close()
    print("감정인식엔진: " + emotion + "이 느껴집니다.")
    return emotions

def getEmotion():
    from openpyxl import load_workbook

    emotion = ''


    #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("/home2/siso/kik/emotionScore.xlsx", data_only=True)
    #시트 이름으로 불러오기
    load_ws = load_wb['Sheet1']

    scorelist = []
    num = 2

    while load_ws['B'+str(num)].value:
        scorelist.append(load_ws['B'+str(num)].value)
        #print(load_ws['B'+str(num)].value)
        num = num+1

    scorelist1 = []
    num1 = 2
  
    while load_ws['C'+str(num1)].value:
        scorelist1.append(load_ws['C'+str(num1)].value)
        #print(load_ws['C'+str(num1)].value)
        num1 = num1 +1
    avg = sum(scorelist)/(num-2)
    avg1 = sum(scorelist1)/(num1 -2)


    emotionScore = (avg+avg1)/2

    #셀 주소로 값 출력
    #print(emotionScore)

    if emotionScore > 0:
        emotion = "행복"
        print("현재 감정은 " + emotion + " 입니다")

    elif emotionScore == 0:
        emotion = "중립"
        print("현재 감정은 " + emotion + " 입니다")
    else:
        emotion = "우울"
        print("현재 감정은 " + emotion + " 입니다")
        
    inputEmotion(0,0,0,True)
    return emotion


def inputEmotion(method,num,emotion_num,final):
    from openpyxl import load_workbook

    write_wb = load_workbook("/home2/siso/kik/emotionScore.xlsx",data_only=False )
    write_ws = write_wb['Sheet1']
    #이름이 있는 시트를 생성
    #write_ws = write_wb.create_sheet('생성시트')

    if final == False:
        if method == 'txt':
            write_ws['B'+str(num+1)] = emotion_num
        else:
            write_ws['C'+str(num+1)] = emotion_num
    else:
        write_ws.delete_cols(2)
        write_ws.insert_cols(2)
        write_ws.delete_cols(3)
        write_ws.insert_cols(3)


    write_wb.save('/home2/siso/kik/emotionScore.xlsx')


def inputEmotion2(data):
    from openpyxl import load_workbook

    write_wb = load_workbook("/home2/siso/kik/emotionScore.xlsx",data_only=False)
    write_ws = write_wb['Sheet2']
    num = 1
    while write_ws['B'+str(num)].value:
        num = num+1
    write_ws['A'+str(num)] = str(num)+'회차'
    write_ws['B'+str(num)] = data
    write_wb.save('/home2/siso/kik/emotionScore.xlsx')


def main():
    global end
    num = 1
    print("준비완료")
    while True:
        recvend()
        print("대기중")
        if end == "false":
            img()
            inputEmotion('img' ,num ,getCNN(num) ,False)

            txt_data = txt()
            inputEmotion('txt',num ,predict(txt_data), False)
            print("내용: " + txt_data)
        else:
            print("====전달 끝=====")
            emotion = getEmotion()
            inputEmotion2(emotion)
            sendEmotion(emotion)
            break
        num = num+1


if __name__ == '__main__':
    device = torch.device("cuda:0" if torch.cuda.is_available() else  "cpu")
    bertmodel, vocab = get_pytorch_kobert_model()
    os.chdir('/home2/siso/Documents/Dkobert/models/')
    torch.cuda.empty_cache()
    model1 = torch.load('7emotions_model.pt')
    torch.cuda.empty_cache()
    tokenizer = get_tokenizer()
    tok = nlp.data.BERTSPTokenizer(tokenizer, vocab, lower=False)
    main()

